"""
Export Module - PDF & Excel Report Export
Milestone 4 - Day 2
Exports financial reports to PDF and Excel formats.
"""
import os
from flask import Blueprint, request, session, send_file, flash, redirect, url_for
from utils.db import query_db
from utils.helpers import login_required, get_current_month, get_current_year, get_month_name, calculate_percentage
from config import Config

# Lazy-load openpyxl styles so missing dependency doesn't break import
try:
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    OPENPYXL_AVAILABLE = True
except ImportError:
    OPENPYXL_AVAILABLE = False

bp = Blueprint('export', __name__, url_prefix='/export')


def _ensure_export_dir():
    """Ensure export directory exists."""
    export_dir = Config.EXPORT_FOLDER
    os.makedirs(export_dir, exist_ok=True)
    return export_dir


@bp.route('/pdf/<report_type>')
@login_required
def export_pdf(report_type):
    """Export report to PDF."""
    user_id = session['user_id']
    month = int(request.args.get('month', get_current_month()))
    year = int(request.args.get('year', get_current_year()))
    month_str = f"{year}-{month:02d}"

    try:
        from fpdf import FPDF
    except ImportError:
        flash('PDF export library not installed. Run: pip install fpdf2', 'danger')
        return redirect(url_for('reports.index'))

    export_dir = _ensure_export_dir()
    filename = f"{report_type}_report_{year}_{month:02d}_{user_id}.pdf"
    filepath = os.path.join(export_dir, filename)

    pdf = FPDF()
    pdf.add_page()

    # Header
    pdf.set_font('Helvetica', 'B', 18)
    pdf.set_text_color(22, 163, 74)  # Green
    pdf.cell(0, 12, 'SMART FINANCE INSIGHTS', ln=True, align='C')
    pdf.set_font('Helvetica', '', 12)
    pdf.set_text_color(100, 100, 100)
    title_map = {
        'expense': f'Monthly Expense Report - {get_month_name(month)} {year}',
        'budget': f'Budget Utilization Report - {get_month_name(month)} {year}',
        'investment': 'Investment Performance Report',
        'goal': 'Financial Goal Progress Report'
    }
    pdf.cell(0, 8, title_map.get(report_type, 'Report'), ln=True, align='C')
    pdf.ln(5)
    pdf.set_draw_color(22, 163, 74)
    pdf.line(10, pdf.get_y(), 200, pdf.get_y())
    pdf.ln(5)

    if report_type == 'expense':
        _pdf_expense_report(pdf, user_id, month_str)
    elif report_type == 'budget':
        _pdf_budget_report(pdf, user_id, month, year, month_str)
    elif report_type == 'investment':
        _pdf_investment_report(pdf, user_id)
    elif report_type == 'goal':
        _pdf_goal_report(pdf, user_id)

    # Footer
    pdf.ln(5)
    pdf.set_font('Helvetica', 'I', 8)
    pdf.set_text_color(150, 150, 150)
    pdf.cell(0, 5, f'Generated on {pdf.get_y():.0f} | Smart Finance Insights', ln=True, align='C')

    pdf.output(filepath)
    flash('PDF report exported successfully!', 'success')
    return send_file(filepath, as_attachment=True, download_name=filename)


def _pdf_expense_report(pdf, user_id, month_str):
    """Build expense section of PDF."""
    pdf.set_font('Helvetica', 'B', 13)
    pdf.set_text_color(0, 0, 0)
    pdf.cell(0, 8, 'Category Summary', ln=True)

    cat_summary = query_db(
        """SELECT category, SUM(amount) as total, COUNT(*) as count FROM expenses
           WHERE user_id=? AND strftime('%Y-%m', date)=?
           GROUP BY category ORDER BY total DESC""",
        (user_id, month_str)
    )
    total = sum(c['total'] for c in cat_summary) if cat_summary else 0

    # Table header
    pdf.set_font('Helvetica', 'B', 10)
    pdf.set_fill_color(22, 163, 74)
    pdf.set_text_color(255, 255, 255)
    pdf.cell(60, 8, 'Category', border=1, fill=True)
    pdf.cell(30, 8, 'Count', border=1, fill=True, align='C')
    pdf.cell(40, 8, 'Amount', border=1, fill=True, align='R')
    pdf.cell(30, 8, 'Percent', border=1, fill=True, align='R', ln=True)

    pdf.set_font('Helvetica', '', 10)
    pdf.set_text_color(0, 0, 0)
    for c in (cat_summary or []):
        pdf.cell(60, 7, c['category'], border=1)
        pdf.cell(30, 7, str(c['count']), border=1, align='C')
        pdf.cell(40, 7, f"Rs. {c['total']:,.0f}", border=1, align='R')
        pdf.cell(30, 7, f"{calculate_percentage(c['total'], total):.1f}%", border=1, align='R', ln=True)

    pdf.set_font('Helvetica', 'B', 10)
    pdf.set_fill_color(240, 240, 240)
    pdf.cell(90, 8, 'TOTAL', border=1, fill=True)
    pdf.cell(40, 8, f"Rs. {total:,.0f}", border=1, fill=True, align='R')
    pdf.cell(30, 8, '100%', border=1, fill=True, align='R', ln=True)


def _pdf_budget_report(pdf, user_id, month, year, month_str):
    """Build budget section of PDF."""
    pdf.set_font('Helvetica', 'B', 13)
    pdf.set_text_color(0, 0, 0)
    pdf.cell(0, 8, 'Budget Utilization', ln=True)

    budgets = query_db("SELECT * FROM budgets WHERE user_id=? AND month=? AND year=?", (user_id, month, year))
    pdf.set_font('Helvetica', 'B', 10)
    pdf.set_fill_color(22, 163, 74)
    pdf.set_text_color(255, 255, 255)
    pdf.cell(45, 8, 'Category', border=1, fill=True)
    pdf.cell(35, 8, 'Budget', border=1, fill=True, align='R')
    pdf.cell(35, 8, 'Spent', border=1, fill=True, align='R')
    pdf.cell(35, 8, 'Remaining', border=1, fill=True, align='R')
    pdf.cell(20, 8, 'Used%', border=1, fill=True, align='R', ln=True)

    pdf.set_font('Helvetica', '', 10)
    pdf.set_text_color(0, 0, 0)
    total_b = total_s = 0
    for b in (budgets or []):
        spent_row = query_db(
            "SELECT COALESCE(SUM(amount),0) as t FROM expenses WHERE user_id=? AND category=? AND strftime('%Y-%m', date)=?",
            (user_id, b['category'], month_str), one=True
        )
        spent = spent_row['t'] if spent_row else 0
        pdf.cell(45, 7, b['category'], border=1)
        pdf.cell(35, 7, f"Rs. {b['amount']:,.0f}", border=1, align='R')
        pdf.cell(35, 7, f"Rs. {spent:,.0f}", border=1, align='R')
        pdf.cell(35, 7, f"Rs. {b['amount']-spent:,.0f}", border=1, align='R')
        pdf.cell(20, 7, f"{calculate_percentage(spent, b['amount']):.0f}%", border=1, align='R', ln=True)
        total_b += b['amount']
        total_s += spent

    pdf.set_font('Helvetica', 'B', 10)
    pdf.set_fill_color(240, 240, 240)
    pdf.cell(45, 8, 'TOTAL', border=1, fill=True)
    pdf.cell(35, 8, f"Rs. {total_b:,.0f}", border=1, fill=True, align='R')
    pdf.cell(35, 8, f"Rs. {total_s:,.0f}", border=1, fill=True, align='R')
    pdf.cell(35, 8, f"Rs. {total_b-total_s:,.0f}", border=1, fill=True, align='R')
    pdf.cell(20, 8, f"{calculate_percentage(total_s, total_b):.0f}%", border=1, fill=True, align='R', ln=True)


def _pdf_investment_report(pdf, user_id):
    """Build investment section of PDF."""
    pdf.set_font('Helvetica', 'B', 13)
    pdf.set_text_color(0, 0, 0)
    pdf.cell(0, 8, 'Investment Performance', ln=True)

    investments = query_db("SELECT * FROM investments WHERE user_id=?", (user_id,))
    pdf.set_font('Helvetica', 'B', 10)
    pdf.set_fill_color(22, 163, 74)
    pdf.set_text_color(255, 255, 255)
    pdf.cell(50, 8, 'Investment', border=1, fill=True)
    pdf.cell(35, 8, 'Asset Type', border=1, fill=True)
    pdf.cell(30, 8, 'Invested', border=1, fill=True, align='R')
    pdf.cell(30, 8, 'Current', border=1, fill=True, align='R')
    pdf.cell(25, 8, 'ROI', border=1, fill=True, align='R', ln=True)

    pdf.set_font('Helvetica', '', 10)
    pdf.set_text_color(0, 0, 0)
    total_inv = total_cur = 0
    for inv in (investments or []):
        pl = inv['current_value'] - inv['invested_amount']
        roi = calculate_percentage(pl, inv['invested_amount'])
        pdf.cell(50, 7, inv['investment_name'][:25], border=1)
        pdf.cell(35, 7, inv['asset_type'], border=1)
        pdf.cell(30, 7, f"Rs. {inv['invested_amount']:,.0f}", border=1, align='R')
        pdf.cell(30, 7, f"Rs. {inv['current_value']:,.0f}", border=1, align='R')
        pdf.cell(25, 7, f"{roi:+.1f}%", border=1, align='R', ln=True)
        total_inv += inv['invested_amount']
        total_cur += inv['current_value']

    pdf.set_font('Helvetica', 'B', 10)
    pdf.set_fill_color(240, 240, 240)
    pdf.cell(85, 8, 'TOTAL', border=1, fill=True)
    pdf.cell(30, 8, f"Rs. {total_inv:,.0f}", border=1, fill=True, align='R')
    pdf.cell(30, 8, f"Rs. {total_cur:,.0f}", border=1, fill=True, align='R')
    pdf.cell(25, 8, f"{calculate_percentage(total_cur-total_inv, total_inv):+.1f}%", border=1, fill=True, align='R', ln=True)


def _pdf_goal_report(pdf, user_id):
    """Build goal section of PDF."""
    pdf.set_font('Helvetica', 'B', 13)
    pdf.set_text_color(0, 0, 0)
    pdf.cell(0, 8, 'Financial Goal Progress', ln=True)

    goals = query_db("SELECT * FROM goals WHERE user_id=? ORDER BY target_date", (user_id,))
    pdf.set_font('Helvetica', 'B', 10)
    pdf.set_fill_color(22, 163, 74)
    pdf.set_text_color(255, 255, 255)
    pdf.cell(55, 8, 'Goal', border=1, fill=True)
    pdf.cell(35, 8, 'Target', border=1, fill=True, align='R')
    pdf.cell(35, 8, 'Saved', border=1, fill=True, align='R')
    pdf.cell(25, 8, 'Remaining', border=1, fill=True, align='R')
    pdf.cell(20, 8, 'Progress', border=1, fill=True, align='R', ln=True)

    pdf.set_font('Helvetica', '', 10)
    pdf.set_text_color(0, 0, 0)
    total_t = total_s = 0
    for g in (goals or []):
        pct = calculate_percentage(g['saved_amount'], g['target_amount'])
        pdf.cell(55, 7, g['goal_name'][:28], border=1)
        pdf.cell(35, 7, f"Rs. {g['target_amount']:,.0f}", border=1, align='R')
        pdf.cell(35, 7, f"Rs. {g['saved_amount']:,.0f}", border=1, align='R')
        pdf.cell(25, 7, f"Rs. {g['target_amount']-g['saved_amount']:,.0f}", border=1, align='R')
        pdf.cell(20, 7, f"{pct:.0f}%", border=1, align='R', ln=True)
        total_t += g['target_amount']
        total_s += g['saved_amount']

    pdf.set_font('Helvetica', 'B', 10)
    pdf.set_fill_color(240, 240, 240)
    pdf.cell(55, 8, 'TOTAL', border=1, fill=True)
    pdf.cell(35, 8, f"Rs. {total_t:,.0f}", border=1, fill=True, align='R')
    pdf.cell(35, 8, f"Rs. {total_s:,.0f}", border=1, fill=True, align='R')
    pdf.cell(25, 8, f"Rs. {total_t-total_s:,.0f}", border=1, fill=True, align='R')
    pdf.cell(20, 8, f"{calculate_percentage(total_s, total_t):.0f}%", border=1, fill=True, align='R', ln=True)


@bp.route('/excel/<report_type>')
@login_required
def export_excel(report_type):
    """Export report to Excel."""
    user_id = session['user_id']
    month = int(request.args.get('month', get_current_month()))
    year = int(request.args.get('year', get_current_year()))
    month_str = f"{year}-{month:02d}"

    try:
        import openpyxl
    except ImportError:
        flash('Excel export library not installed. Run: pip install openpyxl', 'danger')
        return redirect(url_for('reports.index'))

    export_dir = _ensure_export_dir()
    filename = f"{report_type}_report_{year}_{month:02d}_{user_id}.xlsx"
    filepath = os.path.join(export_dir, filename)

    wb = openpyxl.Workbook()
    ws = wb.active
    title_map = {
        'expense': f'Monthly Expense Report - {get_month_name(month)} {year}',
        'budget': f'Budget Utilization Report - {get_month_name(month)} {year}',
        'investment': 'Investment Performance Report',
        'goal': 'Financial Goal Progress Report'
    }
    ws.title = title_map.get(report_type, 'Report')[:31]

    # Styles
    header_font = Font(bold=True, color='FFFFFF', size=12)
    header_fill = PatternFill(start_color='16A34A', end_color='16A34A', fill_type='solid')
    title_font = Font(bold=True, size=16, color='16A34A')
    total_font = Font(bold=True)
    total_fill = PatternFill(start_color='F0F0F0', end_color='F0F0F0', fill_type='solid')
    thin_border = Border(
        left=Side(style='thin'), right=Side(style='thin'),
        top=Side(style='thin'), bottom=Side(style='thin')
    )

    # Title
    ws.merge_cells('A1:E1')
    ws['A1'] = 'SMART FINANCE INSIGHTS'
    ws['A1'].font = title_font
    ws['A1'].alignment = Alignment(horizontal='center')
    ws.merge_cells('A2:E2')
    ws['A2'] = title_map.get(report_type, 'Report')
    ws['A2'].font = Font(size=12, italic=True)
    ws['A2'].alignment = Alignment(horizontal='center')

    row = 4

    if report_type == 'expense':
        row = _excel_expense(ws, row, user_id, month_str, header_font, header_fill, total_font, total_fill, thin_border)
    elif report_type == 'budget':
        row = _excel_budget(ws, row, user_id, month, year, month_str, header_font, header_fill, total_font, total_fill, thin_border)
    elif report_type == 'investment':
        row = _excel_investment(ws, row, user_id, header_font, header_fill, total_font, total_fill, thin_border)
    elif report_type == 'goal':
        row = _excel_goal(ws, row, user_id, header_font, header_fill, total_font, total_fill, thin_border)

    # Auto-adjust column widths (handle merged cells safely)
    from openpyxl.utils import get_column_letter
    max_col = ws.max_column
    for col_idx in range(1, max_col + 1):
        col_letter = get_column_letter(col_idx)
        max_len = 0
        for row_idx in range(1, ws.max_row + 1):
            cell = ws.cell(row=row_idx, column=col_idx)
            try:
                if cell.value:
                    max_len = max(max_len, len(str(cell.value)))
            except Exception:
                pass
        ws.column_dimensions[col_letter].width = min(max_len + 4, 35)

    wb.save(filepath)
    flash('Excel report exported successfully!', 'success')
    return send_file(filepath, as_attachment=True, download_name=filename)


def _excel_expense(ws, row, user_id, month_str, hf, hfill, tf, tfill, border):
    cat_summary = query_db(
        """SELECT category, SUM(amount) as total, COUNT(*) as count FROM expenses
           WHERE user_id=? AND strftime('%Y-%m', date)=?
           GROUP BY category ORDER BY total DESC""",
        (user_id, month_str)
    )
    headers = ['Category', 'Transaction Count', 'Amount (Rs.)', 'Percentage']
    for c, h in enumerate(headers, 1):
        cell = ws.cell(row=row, column=c, value=h)
        cell.font = hf
        cell.fill = hfill
        cell.alignment = Alignment(horizontal='center')
        cell.border = border
    row += 1

    total = sum(c['total'] for c in cat_summary) if cat_summary else 0
    for c in (cat_summary or []):
        ws.cell(row=row, column=1, value=c['category']).border = border
        ws.cell(row=row, column=2, value=c['count']).border = border
        ws.cell(row=row, column=3, value=float(c['total'])).border = border
        ws.cell(row=row, column=4, value=f"{calculate_percentage(c['total'], total):.1f}%").border = border
        row += 1

    ws.cell(row=row, column=1, value='TOTAL').font = tf
    ws.cell(row=row, column=1).fill = tfill
    ws.cell(row=row, column=3, value=float(total)).font = tf
    ws.cell(row=row, column=3).fill = tfill
    ws.cell(row=row, column=4, value='100%').font = tf
    ws.cell(row=row, column=4).fill = tfill
    return row + 1


def _excel_budget(ws, row, user_id, month, year, month_str, hf, hfill, tf, tfill, border):
    budgets = query_db("SELECT * FROM budgets WHERE user_id=? AND month=? AND year=?", (user_id, month, year))
    headers = ['Category', 'Budget (Rs.)', 'Spent (Rs.)', 'Remaining (Rs.)', 'Used %']
    for c, h in enumerate(headers, 1):
        cell = ws.cell(row=row, column=c, value=h)
        cell.font = hf
        cell.fill = hfill
        cell.alignment = Alignment(horizontal='center')
        cell.border = border
    row += 1

    total_b = total_s = 0
    for b in (budgets or []):
        spent_row = query_db(
            "SELECT COALESCE(SUM(amount),0) as t FROM expenses WHERE user_id=? AND category=? AND strftime('%Y-%m', date)=?",
            (user_id, b['category'], month_str), one=True
        )
        spent = spent_row['t'] if spent_row else 0
        ws.cell(row=row, column=1, value=b['category']).border = border
        ws.cell(row=row, column=2, value=float(b['amount'])).border = border
        ws.cell(row=row, column=3, value=float(spent)).border = border
        ws.cell(row=row, column=4, value=float(b['amount'] - spent)).border = border
        ws.cell(row=row, column=5, value=f"{calculate_percentage(spent, b['amount']):.0f}%").border = border
        total_b += b['amount']
        total_s += spent
        row += 1

    ws.cell(row=row, column=1, value='TOTAL').font = tf
    ws.cell(row=row, column=2, value=float(total_b)).font = tf
    ws.cell(row=row, column=3, value=float(total_s)).font = tf
    ws.cell(row=row, column=4, value=float(total_b - total_s)).font = tf
    ws.cell(row=row, column=5, value=f"{calculate_percentage(total_s, total_b):.0f}%").font = tf
    for c in range(1, 6):
        ws.cell(row=row, column=c).fill = tfill
    return row + 1


def _excel_investment(ws, row, user_id, hf, hfill, tf, tfill, border):
    investments = query_db("SELECT * FROM investments WHERE user_id=?", (user_id,))
    headers = ['Investment Name', 'Asset Type', 'Invested (Rs.)', 'Current Value (Rs.)', 'P/L (Rs.)', 'ROI %']
    for c, h in enumerate(headers, 1):
        cell = ws.cell(row=row, column=c, value=h)
        cell.font = hf
        cell.fill = hfill
        cell.alignment = Alignment(horizontal='center')
        cell.border = border
    row += 1

    total_inv = total_cur = 0
    for inv in (investments or []):
        pl = inv['current_value'] - inv['invested_amount']
        roi = calculate_percentage(pl, inv['invested_amount'])
        ws.cell(row=row, column=1, value=inv['investment_name']).border = border
        ws.cell(row=row, column=2, value=inv['asset_type']).border = border
        ws.cell(row=row, column=3, value=float(inv['invested_amount'])).border = border
        ws.cell(row=row, column=4, value=float(inv['current_value'])).border = border
        ws.cell(row=row, column=5, value=float(pl)).border = border
        ws.cell(row=row, column=6, value=f"{roi:+.1f}%").border = border
        total_inv += inv['invested_amount']
        total_cur += inv['current_value']
        row += 1

    ws.cell(row=row, column=1, value='TOTAL').font = tf
    ws.cell(row=row, column=3, value=float(total_inv)).font = tf
    ws.cell(row=row, column=4, value=float(total_cur)).font = tf
    ws.cell(row=row, column=5, value=float(total_cur - total_inv)).font = tf
    ws.cell(row=row, column=6, value=f"{calculate_percentage(total_cur-total_inv, total_inv):+.1f}%").font = tf
    for c in range(1, 7):
        ws.cell(row=row, column=c).fill = tfill
    return row + 1


def _excel_goal(ws, row, user_id, hf, hfill, tf, tfill, border):
    goals = query_db("SELECT * FROM goals WHERE user_id=? ORDER BY target_date", (user_id,))
    headers = ['Goal Name', 'Category', 'Target (Rs.)', 'Saved (Rs.)', 'Remaining (Rs.)', 'Progress %']
    for c, h in enumerate(headers, 1):
        cell = ws.cell(row=row, column=c, value=h)
        cell.font = hf
        cell.fill = hfill
        cell.alignment = Alignment(horizontal='center')
        cell.border = border
    row += 1

    total_t = total_s = 0
    for g in (goals or []):
        pct = calculate_percentage(g['saved_amount'], g['target_amount'])
        ws.cell(row=row, column=1, value=g['goal_name']).border = border
        ws.cell(row=row, column=2, value=g['category']).border = border
        ws.cell(row=row, column=3, value=float(g['target_amount'])).border = border
        ws.cell(row=row, column=4, value=float(g['saved_amount'])).border = border
        ws.cell(row=row, column=5, value=float(g['target_amount'] - g['saved_amount'])).border = border
        ws.cell(row=row, column=6, value=f"{pct:.0f}%").border = border
        total_t += g['target_amount']
        total_s += g['saved_amount']
        row += 1

    ws.cell(row=row, column=1, value='TOTAL').font = tf
    ws.cell(row=row, column=3, value=float(total_t)).font = tf
    ws.cell(row=row, column=4, value=float(total_s)).font = tf
    ws.cell(row=row, column=5, value=float(total_t - total_s)).font = tf
    ws.cell(row=row, column=6, value=f"{calculate_percentage(total_s, total_t):.0f}%").font = tf
    for c in range(1, 7):
        ws.cell(row=row, column=c).fill = tfill
    return row + 1
